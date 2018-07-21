# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import json, csv, os
from scrapy.exporters import CsvItemExporter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Color
from scrapy.exceptions import DropItem

# stage 1: read in from interested cars file and car_filter.txt out cars
# if the car title does not have any strings from the filter, remove the result
# if the car's price is outside of the price filter range, remove the car
class CarawlerFilterPipeline(object):

    # open filter files if they exist
    def __init__(self):
        self.car_filter_filename = 'carawler/input/car_filter.txt'
        self.price_filter_filename = 'carawler/input/price_filter.txt'
        try:
            with open(self.car_filter_filename, 'rb') as car_filter:
                self.car_filter = car_filter.readlines()
                for i in range(0, len(self.car_filter)):
                    self.car_filter[i] = self.car_filter[i].strip('\n').replace('*', '')
        except:
            self.cars = None
        try:
            with open(self.price_filter_filename, 'rb') as price_filter:
                self.min_price = int(price_filter.readline())
                self.max_price = int(price_filter.readline())
        except:
            self.min_price = 0
            self.max_price = 10**9 # a billion...

    def process_item(self, item, spider):
        ids_to_remove = []
        for id in item:
            try:
                title = item[id]['title'].encode('utf-8').strip().lower()
            except:
                title = None

            # if none of the cars in the car filter are within the post title, remove this post from the results
            if title is not None:
                if any(substring in title for substring in self.car_filter):
                    # if the post is within the filter, check if the price is also in the filter; remove if not
                    try:
                        price = int(item[id]['price'].replace('$',''))
                    except:  # price doesn't exist
                        price = None
                        ids_to_remove.append(id)
                    if price is not None:
                        if self.max_price < price or self.min_price > price:
                            ids_to_remove.append(id)
                else:
                    ids_to_remove.append(id)

            # post has no title so remove this from results
            else:
                ids_to_remove.append(id)

        # remove items that did not get past filters
        for id in ids_to_remove:
            del item[id]

        return item

# filter out duplicates, from Scrapy documentation
# from scrapy.exceptions import DropItem
#
# class DuplicatesPipeline(object):
#
#     def __init__(self):
#         self.ids_seen = set()
#
#     def process_item(self, item, spider):
#         if item['id'] in self.ids_seen:
#             raise DropItem("Duplicate item found: %s" % item)
#         else:
#             self.ids_seen.add(item['id'])
#             return item

# stage 2: store data in csv file
class CarawlerCSVPipeline(object):

    def __init__(self):
        try:
            with open('carawler/output/cars.csv', 'r') as f:
                try:
                    self.cars_csv = f.readlines()
                    for s in xrange(0, len(self.cars_csv)):
                        self.cars_csv[s] = self.cars_csv[s].strip('\r\n').strip(' ')
                        print "-----------------------------self.cars[s]-----------------------------------"
                        print self.cars_csv[s]
                except:
                    self.cars_csv = []
        # file does not already exist
        except:
            with open('carawler/output/cars.csv', 'w'):
                self.cars_csv = []

    def process_item(self, item, spider):
        with open('carawler/output/cars.csv', 'a+') as f:
            self.csvwriter = csv.writer(f, delimiter=',')
            for id in item:
                try:
                    title = item[id]['title'].encode('utf-8').strip().replace(',', '')
                except:
                    title = 'N/A'
                try:
                    price = item[id]['price'].encode('utf-8').strip().replace(',', '')
                except:
                    price = 'N/A'
                try:
                    location = item[id]['location'].encode('utf-8').strip().replace(',', '')
                except:
                    location = 'N/A'
                try:
                    post_time = item[id]['post_time'].encode('utf-8').strip()
                except:
                    post_time = 'N/A'
                try:
                    url = item[id]['url'].encode('utf-8').strip()
                except:
                    url = 'N/A'
                line = ','.join(map(str, [id, title, price, location, post_time, url])).strip()
                # item duplicate filter: if this line isn't in the results already, add it
                if len(self.cars_csv) == 0 or line not in self.cars_csv:
                    self.csvwriter.writerow(line.split(','))

        return item

# stage 3: convert csv to xlsx to store data in more readable format
class CarawlerConvertToSpreadsheetPipeline(object):

    def __init__(self):
        self.csvfilename = 'carawler/output/cars.csv'
        self.spreadsheetfilename = 'carawler/output/cars.xlsx'

    def process_item(self, item, spider):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cars"
        with open(self.csvfilename, 'r') as f:
            reader = csv.reader(f)
            for r, row in enumerate(reader, start=1):
                for c, val in enumerate(row, start=1):
                    # r is row number, c, is column number (both 1-indexed)
                    ws.cell(row=r, column=c).value = val
                    if r == 1: # headers = blue
                        ws.cell(row=r, column=c).font = Font(size=16)
                        ws.cell(row=r, column=c).fill = PatternFill(fgColor=Color('0087BD'),
                                                                   fill_type='solid')
                    else: # odd rows = gray
                        ws.cell(row=r, column=c).font = Font(size=14)
                        if r % 2 == 1: # row is odd but not equal to 1
                            ws.cell(row=r, column=c).fill = PatternFill(fgColor=Color('BFC1C2'),
                                                                        fill_type='solid')
            # adjust column lengths in worksheet based on widest cells per column
            for col in ws.columns:
                max_length = 0
                column = col[0].column  # Get the column name
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2.0) * 1.1
                ws.column_dimensions[column].width = adjusted_width
        wb.save(self.spreadsheetfilename)

        return item